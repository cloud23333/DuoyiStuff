from __future__ import annotations

import csv
import io
import json
from collections.abc import Callable, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .engine import stockout_mask_cut
from .eval_forecast import EvalRow
from .forecast_curve import build_forecast_curve

RECOMMENDATION_FIELDS = [
    ("internal_order_id", "订单号"),
    ("__item_identity__", "商品定位"),
    ("line_order_qty", "订单行数量"),
    ("recommended_ship", "建议发货量"),
    ("decision_reason", "SKU建议"),
    ("order_decision_reason", "订单建议"),
    ("__action_notes__", "处理提示"),
    ("店铺款式编码", "店铺款式编码"),
    ("店铺商品编码", "店铺商品编码"),
    ("原始商品编码", "订单商品编码"),
    ("key_order_qty", "同款同SKU总下单量"),
    ("key_recommended_total", "同款同SKU建议总量"),
    ("gap", "缺口"),
    ("__available_stock__", "可用库存"),
    ("wh", "平台仓内库存"),
    ("pending_recv", "平台待收货库存"),
    ("shipping_in_progress", "发货中数量"),
    ("pending_ship", "平台待发货库存"),
    ("forecast_stocking_period_sales", "预测备货期销量"),
    ("forecast_daily_sales", "预测日均销量"),
    ("effective_daily_sales", "异常调整后日均销量"),
    ("stocking_days", "备货逻辑天数"),
    ("recommended_ship_before_small_change_rule", "30%规则前建议发货量"),
    ("small_change_ratio_before_rule", "30%规则前变动比例"),
    ("small_change_keep_warning", "30%内免改数量提示"),
    ("order_recommended_ship_total_before_threshold", "订单阈值前建议总量"),
    ("min_order_ship_qty_threshold", "最小发货阈值"),
    ("order_low_qty_warning", "订单低于起发量提示"),
    ("min_order_ship_qty_exempt_warning", "小于10不发豁免资格提示"),
    ("min_order_ship_qty_exempt_applied_warning", "小于10不发豁免生效提示"),
    ("forecast_strategy", "预测策略"),
    ("demand_profile", "需求类型"),
    ("anomaly_flags", "异常标记"),
    ("service_level", "服务水平"),
    ("forecast_model", "预测模型"),
    ("sku_forecast_abs_error", "SKU预测绝对误差"),
    ("sku_forecast_signed_error", "SKU预测偏差"),
    ("__plot__", "销量与预测曲线"),
    ("系统商品编码", "系统商品编码"),
    ("sku_code_check", "SKU编码校验"),
    ("base_stock_qty", "保底库存目标"),
    ("base_stock_gap", "保底库存缺口"),
    ("base_stock_triggered_warning", "保底是否触发"),
    ("intercept_reason", "拦截原因"),
    ("order_intercept_warning", "订单拦截导致不发提示"),
]

PLOT_COLUMN_NAME = "销量与预测曲线"
PLOT_COLUMN_WIDTH_CHARS = 46.0
PLOT_ROW_HEIGHT_POINTS = 100.0
PLOT_IMAGE_WIDTH_PX = 320
PLOT_IMAGE_HEIGHT_PX = 128

RECOMMENDATION_FREEZE_PANES = "E2"
HEADER_ROW_HEIGHT_POINTS = 24.0
DEFAULT_ROW_HEIGHT_POINTS = 34.0
MIN_COLUMN_WIDTH_CHARS = 8.0
MAX_COLUMN_WIDTH_CHARS = 34.0
TEXT_WRAP_COLUMN_WIDTH_CHARS = 30.0

HEADER_COMMENTS = {
    "建议发货量": "最终建议执行数量，优先查看或复制此列。",
    "商品定位": "合并显示店铺款式编码、店铺商品编码和订单商品编码，减少冻结列占用。",
    "处理提示": "汇总影响建议量的主要原因和异常提示。",
    "可用库存": "平台仓内库存 + 平台待收货库存 + 发货中数量。",
    "缺口": "预测备货期需求扣减可用库存后的缺口。",
}

INTEGER_FORMAT = "#,##0"
DECIMAL_FORMAT = "#,##0.####"
PERCENT_FORMAT = "0%"

PERCENT_FORMAT_FIELDS = {
    "service_level",
    "small_change_ratio_before_rule",
}

DECIMAL_FORMAT_FIELDS = {
    "effective_daily_sales",
    "forecast_daily_sales",
    "forecast_stocking_period_sales",
    "stocking_days",
    "sku_forecast_abs_error",
    "sku_forecast_signed_error",
    "wh",
    "pending_recv",
    "shipping_in_progress",
    "pending_ship",
    "__available_stock__",
}

TEXT_WRAP_COLUMNS = {
    "商品定位",
    "处理提示",
    "异常标记",
}

HEADER_GROUP_FILLS = {
    "action": PatternFill("solid", fgColor="1F4E78"),
    "stock": PatternFill("solid", fgColor="5B7F55"),
    "rule": PatternFill("solid", fgColor="8064A2"),
    "forecast": PatternFill("solid", fgColor="4F81BD"),
    "diagnostic": PatternFill("solid", fgColor="7F6000"),
}
HEADER_FONT = Font(color="FFFFFF", bold=True)
GRID_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ACTION_FILL = PatternFill("solid", fgColor="FFF2CC")
ACTION_STRONG_FILL = PatternFill("solid", fgColor="FCE4D6")
MUTED_FILL = PatternFill("solid", fgColor="F2F2F2")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
INFO_FILL = PatternFill("solid", fgColor="E2F0D9")
MUTED_FONT = Font(color="808080")

QUALITY_FIELDS = [
    ("type", "问题类型"),
    ("internal_order_id", "内部订单号"),
    ("skc", "店铺款式编码"),
    ("skuid", "店铺商品编码"),
    ("order_sku", "原始商品编码"),
    ("system_sku", "系统商品编码"),
    ("message", "问题说明"),
]

DECISION_REASON_MAP = {
    "ship_all": "全发",
    "ship_partial": "部分发",
    "hold": "暂不发",
}

INTERCEPT_REASON_MAP = {
    "skc": "命中SKC拦截",
    "skuid": "命中SKUID拦截",
    "skc_and_skuid": "命中SKC和SKUID拦截",
}

SKU_CHECK_MAP = {
    "exact_match": "完全一致",
    "normalized_match": "标准化一致",
    "diff": "不一致",
    "missing_key": "缺少销售匹配",
}

QUALITY_TYPE_MAP = {
    "sku_code_diff": "SKU编码不一致",
    "missing_sales_key": "缺少销售匹配",
    "missing_daily_sales": "缺少每日销量数据",
}

QUALITY_MESSAGE_MAP = {
    "Order SKU and system SKU differ after normalization": "原始商品编码与系统商品编码在标准化后仍不一致",
    "No sales row found for (SKC, SKUID)": "未找到对应销售键 (SKC, SKUID)",
    "Missing daily sales data for (SKC, SKUID)": "未找到对应每日销量数据 (SKC, SKUID)",
}

INT_FORMAT_FIELDS = {
    "line_order_qty",
    "key_order_qty",
    "gap",
    "base_stock_qty",
    "base_stock_gap",
    "key_recommended_total",
    "recommended_ship_before_small_change_rule",
    "recommended_ship",
    "order_recommended_ship_total_before_threshold",
    "min_order_ship_qty_threshold",
}

FORECAST_ERROR_FIELDS = {"sku_forecast_abs_error", "sku_forecast_signed_error"}

SUMMARY_INT_FORMAT_FIELDS = {
    "order_lines",
    "sales_rows",
    "matched_order_lines",
    "total_order_qty",
    "total_recommended_qty",
    "small_change_kept_lines",
    "base_stock_qty",
    "base_stock_triggered_skus",
    "base_stock_triggered_lines",
    "quality_issue_rows",
    "duplicate_sales_keys",
    "min_order_ship_qty_threshold",
    "low_qty_orders_before_exempt",
    "low_qty_order_lines_before_exempt",
    "low_qty_orders",
    "low_qty_order_lines",
    "low_qty_orders_exempted",
    "low_qty_order_lines_exempted",
    "sku_order_limit_rule_count",
    "sku_order_limit_capped_lines",
    "excluded_skc_rule_count",
    "excluded_skuid_rule_count",
    "intercepted_order_lines",
    "intercepted_orders",
}

DECISION_FIELDS = {"decision_reason", "order_decision_reason"}

WARNING_FIELDS = {
    "order_low_qty_warning",
    "min_order_ship_qty_exempt_warning",
    "min_order_ship_qty_exempt_applied_warning",
    "order_intercept_warning",
    "small_change_keep_warning",
    "base_stock_triggered_warning",
}

SUMMARY_FIELDS = [
    ("order_lines", "订单行数"),
    ("sales_rows", "销售行数"),
    ("matched_order_lines", "匹配订单行数"),
    ("join_coverage_pct", "匹配覆盖率_百分比"),
    ("total_order_qty", "总下单量"),
    ("total_recommended_qty", "建议发货总量"),
    ("small_change_kept_lines", "触发30%免改行数"),
    ("decision_ship_all", "建议_全发_行数"),
    ("decision_ship_partial", "建议_部分发_行数"),
    ("decision_hold", "建议_暂不发_行数"),
    ("sku_check_exact_match", "SKU校验_完全一致_行数"),
    ("sku_check_normalized_match", "SKU校验_标准化一致_行数"),
    ("sku_check_diff", "SKU校验_不一致_行数"),
    ("sku_check_missing_key", "SKU校验_缺少销售匹配_行数"),
    ("quality_issue_rows", "质量问题行数"),
    ("duplicate_sales_keys", "销售重复键数量"),
    ("demand_profile_summary", "需求类型分布"),
    ("anomaly_flag_summary", "异常标记分布"),
    ("service_level_summary", "服务水平分布"),
    ("forecast_model_summary", "预测模型分布"),
    ("全局服务水平偏移", "全局服务水平偏移"),
    ("base_stock_qty", "保底库存目标"),
    ("base_stock_triggered_skus", "保底触发SKU数"),
    ("base_stock_triggered_lines", "保底触发订单行数"),
    ("min_order_ship_qty_threshold", "最小发货阈值"),
    ("low_qty_orders_before_exempt", "阈值前低发货量订单数"),
    ("low_qty_order_lines_before_exempt", "阈值前低发货量订单行数"),
    ("low_qty_orders", "低于阈值订单数_提示"),
    ("low_qty_order_lines", "低于阈值订单行数_提示"),
    ("low_qty_orders_exempted", "低于阈值豁免订单数"),
    ("low_qty_order_lines_exempted", "低于阈值豁免订单行数"),
    ("sku_order_limit_rule_count", "订单内SKU限额规则数"),
    ("sku_order_limit_capped_lines", "触发订单内SKU限额行数"),
    ("excluded_skc_rule_count", "SKC拦截规则数"),
    ("excluded_skuid_rule_count", "SKUID拦截规则数"),
    ("intercepted_order_lines", "命中拦截订单行数"),
    ("intercepted_orders", "拦截导致不发订单数"),
]

FORECAST_EVAL_SUMMARY_KEY = "预测回测评估"
FORECAST_EVAL_SUMMARY_FIELDS = [
    ("evaluated_skus", "评估SKU数"),
    ("skipped_insufficient_history", "历史不足跳过SKU数"),
    ("holdout_days", "留出天数"),
    ("mae", "MAE"),
    ("wape", "WAPE"),
    ("bias", "平均偏差"),
]
FORECAST_EVAL_STRATEGY_FIELDS = [
    ("count", "SKU数"),
    ("mae", "MAE"),
    ("wape", "WAPE"),
    ("bias", "平均偏差"),
]
FORECAST_EVAL_INT_FIELDS = {
    "evaluated_skus",
    "skipped_insufficient_history",
    "holdout_days",
    "count",
}


def export_reports(
    out_dir: str | Path,
    recommendations: list[dict[str, object]],
    quality_rows: list[dict[str, object]],
    summary: dict[str, object],
    daily_sales_by_key: dict[tuple[str, str], tuple[int, ...]],
    eval_rows: Sequence[EvalRow] | None = None,
    eval_summary: dict[str, object] | None = None,
) -> dict[str, Path]:
    output_dir = Path(out_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    recommendation_path = output_dir / "发货建议明细.xlsx"
    quality_path = output_dir / "数据质量报告.csv"
    summary_path = output_dir / "运行摘要.json"
    _remove_stale_recommendation_csv(output_dir / "发货建议明细.csv")

    quality_columns = [target for _, target in QUALITY_FIELDS]

    _annotate_forecast_error(recommendations, eval_rows or ())
    plot_cache = _build_plot_cache(recommendations, daily_sales_by_key)
    localized_recommendations = [_localize_recommendation_row(row) for row in recommendations]
    localized_quality_rows = [_localize_quality_row(row) for row in quality_rows]
    localized_summary = _localize_summary(summary, eval_summary=eval_summary)

    _write_recommendations_xlsx(
        path=recommendation_path,
        rows=localized_recommendations,
        recommendation_rows=recommendations,
        plot_cache=plot_cache,
    )
    _write_csv(quality_path, localized_quality_rows, quality_columns)
    summary_path.write_text(
        json.dumps(localized_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return {
        "final_recommendation": recommendation_path,
        "quality_report": quality_path,
        "run_summary": summary_path,
    }


def _remove_stale_recommendation_csv(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        return


def _annotate_forecast_error(
    recommendations: list[dict[str, object]],
    eval_rows: Sequence[EvalRow],
) -> None:
    error_by_key: dict[tuple[str, str], EvalRow] = {
        (row.skc, row.skuid): row for row in eval_rows
    }
    for row in recommendations:
        key = (str(row.get("店铺款式编码", "")), str(row.get("店铺商品编码", "")))
        eval_row = error_by_key.get(key)
        if eval_row is None:
            row["sku_forecast_abs_error"] = ""
            row["sku_forecast_signed_error"] = ""
            continue
        row["sku_forecast_abs_error"] = eval_row.abs_error
        row["sku_forecast_signed_error"] = eval_row.signed_error


def _build_plot_cache(
    recommendations: list[dict[str, object]],
    daily_sales_by_key: dict[tuple[str, str], tuple[int, ...]],
) -> dict[tuple[str, str], bytes]:
    """Render one plot for every unique SKU in the recommendation rows."""
    from .plots import render_sku_plot

    cache: dict[tuple[str, str], bytes] = {}
    plot_keys = _plot_keys(recommendations)
    for row in recommendations:
        key = (str(row.get("店铺款式编码", "")), str(row.get("店铺商品编码", "")))
        if key in cache or key not in plot_keys:
            continue
        history = daily_sales_by_key.get(key, ())
        forecast = build_forecast_curve(
            strategy=str(row.get("forecast_strategy", "") or ""),
            forecast_daily_sales=_as_float(row.get("forecast_daily_sales", 0.0)),
            stocking_days=_as_float(row.get("stocking_days", 0.0)),
        )
        if not history and not forecast:
            continue
        masked_tail_from = stockout_mask_cut(
            daily_sales=history,
            stock_in_warehouse=_as_float(row.get("wh", 0.0)),
        )
        title = _plot_title(row, key)
        cache[key] = render_sku_plot(
            history=history,
            forecast=forecast,
            title=title,
            masked_tail_from=masked_tail_from,
            forecast_daily_sales=_as_float(row.get("forecast_daily_sales", 0.0)),
        )
    return cache


def _plot_title(
    row: dict[str, object],
    key: tuple[str, str],
) -> str:
    base = f"SKC:{key[0]} / SKUID:{key[1]}"
    model = str(row.get("forecast_model", "") or "").strip()
    if not model:
        return base
    return f"{base} | {model}"


def _plot_keys(
    recommendations: list[dict[str, object]],
) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    for row in recommendations:
        key = (str(row.get("店铺款式编码", "")), str(row.get("店铺商品编码", "")))
        keys.add(key)
    return keys


def _as_float(value: object) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return 0.0


def _write_recommendations_xlsx(
    *,
    path: Path,
    rows: list[dict[str, object]],
    recommendation_rows: list[dict[str, object]],
    plot_cache: dict[tuple[str, str], bytes],
) -> None:
    fields = RECOMMENDATION_FIELDS
    columns = [target for _, target in fields]
    plot_col_idx = columns.index(PLOT_COLUMN_NAME) + 1
    plot_col_letter = get_column_letter(plot_col_idx)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "发货建议明细"

    for col_idx, name in enumerate(columns, start=1):
        worksheet.cell(row=1, column=col_idx, value=name)

    worksheet.column_dimensions[plot_col_letter].width = PLOT_COLUMN_WIDTH_CHARS

    for row_offset, (localized, source) in enumerate(
        zip(rows, recommendation_rows, strict=True)
    ):
        excel_row = row_offset + 2
        for col_idx, name in enumerate(columns, start=1):
            if name == PLOT_COLUMN_NAME:
                continue
            worksheet.cell(row=excel_row, column=col_idx, value=localized.get(name, ""))

        key = (str(source.get("店铺款式编码", "")), str(source.get("店铺商品编码", "")))
        image_bytes = plot_cache.get(key)
        if image_bytes is not None:
            image_stream = io.BytesIO(image_bytes)
            excel_image = XlsxImage(image_stream)
            excel_image.width = PLOT_IMAGE_WIDTH_PX
            excel_image.height = PLOT_IMAGE_HEIGHT_PX
            excel_image.anchor = f"{plot_col_letter}{excel_row}"
            worksheet.add_image(excel_image)
            worksheet.row_dimensions[excel_row].height = PLOT_ROW_HEIGHT_POINTS

    _style_recommendations_worksheet(worksheet=worksheet, fields=fields)
    workbook.save(path)


def _style_recommendations_worksheet(
    *,
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    max_row = worksheet.max_row
    max_col = len(fields)
    last_col_letter = get_column_letter(max_col)
    worksheet.freeze_panes = RECOMMENDATION_FREEZE_PANES
    worksheet.auto_filter.ref = f"A1:{last_col_letter}{max_row}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT_POINTS
    worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT_POINTS

    for col_idx, (source, name) in enumerate(fields, start=1):
        header = worksheet.cell(row=1, column=col_idx)
        header.fill = _header_fill_for_source(source)
        header.font = HEADER_FONT
        header.border = GRID_BORDER
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if name in HEADER_COMMENTS:
            header.comment = Comment(HEADER_COMMENTS[name], "shipment-planner")

    for row_idx in range(2, max_row + 1):
        for col_idx, (source, name) in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = _cell_alignment(source, name)
            cell.border = GRID_BORDER
            if source in INT_FORMAT_FIELDS:
                cell.number_format = INTEGER_FORMAT
            elif source in PERCENT_FORMAT_FIELDS:
                cell.number_format = PERCENT_FORMAT
            elif source in DECIMAL_FORMAT_FIELDS:
                cell.number_format = DECIMAL_FORMAT

    _apply_column_widths(worksheet, fields)
    _apply_recommendation_conditional_formatting(worksheet, fields)


def _header_fill_for_source(source: str) -> PatternFill:
    if source in {
        "internal_order_id",
        "__item_identity__",
        "店铺款式编码",
        "店铺商品编码",
        "原始商品编码",
        "line_order_qty",
        "recommended_ship",
        "decision_reason",
        "order_decision_reason",
        "__action_notes__",
    }:
        return HEADER_GROUP_FILLS["action"]
    if source in {
        "key_order_qty",
        "key_recommended_total",
        "gap",
        "__available_stock__",
        "wh",
        "pending_recv",
        "shipping_in_progress",
        "pending_ship",
    }:
        return HEADER_GROUP_FILLS["stock"]
    if source in {
        "recommended_ship_before_small_change_rule",
        "small_change_ratio_before_rule",
        "small_change_keep_warning",
        "order_recommended_ship_total_before_threshold",
        "min_order_ship_qty_threshold",
        "order_low_qty_warning",
        "min_order_ship_qty_exempt_warning",
        "min_order_ship_qty_exempt_applied_warning",
    }:
        return HEADER_GROUP_FILLS["rule"]
    if source in {
        "forecast_stocking_period_sales",
        "forecast_daily_sales",
        "effective_daily_sales",
        "stocking_days",
        "forecast_strategy",
        "demand_profile",
        "anomaly_flags",
        "service_level",
        "forecast_model",
        "sku_forecast_abs_error",
        "sku_forecast_signed_error",
        "__plot__",
    }:
        return HEADER_GROUP_FILLS["forecast"]
    return HEADER_GROUP_FILLS["diagnostic"]


def _cell_alignment(source: str, name: str) -> Alignment:
    return Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=name in TEXT_WRAP_COLUMNS,
    )


def _apply_column_widths(
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    for col_idx, (source, name) in enumerate(fields, start=1):
        letter = get_column_letter(col_idx)
        if name == PLOT_COLUMN_NAME:
            worksheet.column_dimensions[letter].width = PLOT_COLUMN_WIDTH_CHARS
            continue
        width = _recommended_column_width(worksheet, col_idx, source, name)
        worksheet.column_dimensions[letter].width = width


def _recommended_column_width(
    worksheet,
    col_idx: int,
    source: str,
    name: str,
) -> float:
    if name in TEXT_WRAP_COLUMNS:
        return 24.0 if source == "__item_identity__" else TEXT_WRAP_COLUMN_WIDTH_CHARS
    if source == "internal_order_id":
        return 13.0
    if source in {"店铺款式编码", "店铺商品编码"}:
        return 15.0
    if source in {"原始商品编码", "系统商品编码"}:
        return 16.0
    if source in {"decision_reason", "order_decision_reason"}:
        return 8.0
    if source == "line_order_qty":
        return 10.0
    if source == "recommended_ship":
        return 11.0

    max_width = _display_width(name)
    for row_idx in range(2, worksheet.max_row + 1):
        max_width = max(
            max_width,
            _display_width(worksheet.cell(row=row_idx, column=col_idx).value),
        )
    return min(max(max_width + 2, MIN_COLUMN_WIDTH_CHARS), MAX_COLUMN_WIDTH_CHARS)


def _display_width(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    return sum(2 if ord(char) > 127 else 1 for char in text)


def _apply_recommendation_conditional_formatting(
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    if worksheet.max_row < 2:
        return

    columns = {name: get_column_letter(idx) for idx, (_, name) in enumerate(fields, start=1)}
    last_row = worksheet.max_row

    recommended_col = columns["建议发货量"]
    worksheet.conditional_formatting.add(
        f"{recommended_col}2:{recommended_col}{last_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=ACTION_FILL,
            font=Font(bold=True, color="9C6500"),
        ),
    )
    worksheet.conditional_formatting.add(
        f"{recommended_col}2:{recommended_col}{last_row}",
        CellIsRule(
            operator="equal",
            formula=["0"],
            fill=MUTED_FILL,
            font=MUTED_FONT,
        ),
    )

    notes_col = columns["处理提示"]
    worksheet.conditional_formatting.add(
        f"{notes_col}2:{notes_col}{last_row}",
        FormulaRule(formula=[f"LEN(TRIM(${notes_col}2))>0"], fill=WARNING_FILL),
    )

    for name in (
        "30%内免改数量提示",
        "订单低于起发量提示",
        "小于10不发豁免资格提示",
        "小于10不发豁免生效提示",
        "保底是否触发",
        "订单拦截导致不发提示",
    ):
        col = columns[name]
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            FormulaRule(formula=[f'${col}2="是"'], fill=INFO_FILL),
        )

    check_col = columns["SKU编码校验"]
    worksheet.conditional_formatting.add(
        f"{check_col}2:{check_col}{last_row}",
        FormulaRule(
            formula=[f'OR(${check_col}2="不一致",${check_col}2="缺少销售匹配")'],
            fill=WARNING_FILL,
        ),
    )

    intercept_col = columns["拦截原因"]
    worksheet.conditional_formatting.add(
        f"{intercept_col}2:{intercept_col}{last_row}",
        FormulaRule(formula=[f"LEN(TRIM(${intercept_col}2))>0"], fill=ACTION_STRONG_FILL),
    )


def _write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        if rows:
            writer.writerows(rows)


def _localize_recommendation_row(row: dict[str, object]) -> dict[str, object]:
    return _localize_row(
        row=row,
        fields=RECOMMENDATION_FIELDS,
        default_value="",
        value_mapper=_localize_recommendation_value,
        derived_value_mapper=_recommendation_derived_value,
    )


def _localize_quality_row(row: dict[str, object]) -> dict[str, object]:
    return _localize_row(
        row=row,
        fields=QUALITY_FIELDS,
        default_value="",
        value_mapper=_localize_quality_value,
    )


def _localize_row(
    *,
    row: dict[str, object],
    fields: Sequence[tuple[str, str]],
    default_value: object,
    value_mapper: Callable[[str, object], object],
    derived_value_mapper: Callable[[str, dict[str, object]], object] | None = None,
) -> dict[str, object]:
    localized: dict[str, object] = {}
    for source, target in fields:
        if source == "__plot__":
            localized[target] = ""
            continue
        if source.startswith("__") and derived_value_mapper is not None:
            localized[target] = derived_value_mapper(source, row)
            continue
        localized[target] = value_mapper(source, row.get(source, default_value))
    return localized


def _recommendation_derived_value(source: str, row: dict[str, object]) -> object:
    if source == "__item_identity__":
        return _build_item_identity(row)
    if source == "__action_notes__":
        return _build_action_notes(row)
    if source == "__available_stock__":
        return _format_int_like(
            round(
                _as_float(row.get("wh"))
                + _as_float(row.get("pending_recv"))
                + _as_float(row.get("shipping_in_progress")),
                4,
            )
        )
    return ""


def _build_item_identity(row: dict[str, object]) -> str:
    skc = str(row.get("店铺款式编码", "")).strip()
    skuid = str(row.get("店铺商品编码", "")).strip()
    order_sku = str(row.get("原始商品编码", "")).strip()
    parts = [
        f"SKC {skc}" if skc else "",
        f"SKUID {skuid}" if skuid else "",
        f"SKU {order_sku}" if order_sku else "",
    ]
    return "\n".join(part for part in parts if part)


def _build_action_notes(row: dict[str, object]) -> str:
    notes: list[str] = []

    sku_code_check = str(row.get("sku_code_check", "")).strip()
    if sku_code_check == "diff":
        notes.append("SKU不一致")
    elif sku_code_check == "missing_key":
        notes.append("缺少销售匹配")

    intercept_reason = str(row.get("intercept_reason", "")).strip()
    if intercept_reason:
        notes.append(str(INTERCEPT_REASON_MAP.get(intercept_reason, intercept_reason)))

    if _is_yes(row.get("order_intercept_warning")):
        notes.append("订单拦截导致不发")
    if _is_yes(row.get("order_low_qty_warning")):
        notes.append("订单低于起发量")
    if _is_yes(row.get("min_order_ship_qty_exempt_applied_warning")):
        notes.append("小于10豁免生效")
    elif _is_yes(row.get("min_order_ship_qty_exempt_warning")):
        notes.append("小于10豁免候选")
    if _is_yes(row.get("small_change_keep_warning")):
        notes.append("30%内免改")
    if _is_yes(row.get("base_stock_triggered_warning")):
        notes.append("保底触发")

    anomaly_flags = str(row.get("anomaly_flags", "")).strip()
    if anomaly_flags and anomaly_flags != "无":
        notes.append(f"销量异常：{anomaly_flags}")

    return "；".join(notes)


def _is_yes(value: object) -> bool:
    return str(value).strip() in {"yes", "是"}


def _localize_summary(
    summary: dict[str, object],
    *,
    eval_summary: dict[str, object] | None = None,
) -> dict[str, object]:
    localized: dict[str, object] = {}
    for source, target in SUMMARY_FIELDS:
        value = summary.get(source, 0)
        if source in SUMMARY_INT_FORMAT_FIELDS:
            localized[target] = _format_int_like(value)
            continue
        localized[target] = value
    if eval_summary is not None:
        localized[FORECAST_EVAL_SUMMARY_KEY] = _localize_eval_summary(eval_summary)
    return localized


def _localize_eval_summary(eval_summary: dict[str, object]) -> dict[str, object]:
    localized: dict[str, object] = {}
    for source, target in FORECAST_EVAL_SUMMARY_FIELDS:
        localized[target] = _localize_eval_summary_value(source, eval_summary.get(source))

    by_strategy = eval_summary.get("by_strategy", {})
    if isinstance(by_strategy, dict):
        localized["按策略"] = {
            str(strategy): _localize_eval_strategy_summary(stats)
            for strategy, stats in by_strategy.items()
            if isinstance(stats, dict)
        }
    else:
        localized["按策略"] = {}
    return localized


def _localize_eval_strategy_summary(stats: dict[object, object]) -> dict[str, object]:
    localized: dict[str, object] = {}
    for source, target in FORECAST_EVAL_STRATEGY_FIELDS:
        localized[target] = _localize_eval_summary_value(source, stats.get(source))
    return localized


def _localize_eval_summary_value(source: str, value: object) -> object:
    if source in FORECAST_EVAL_INT_FIELDS:
        return _format_int_like(value)
    return value


def _format_int_like(value: object) -> object:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _localize_recommendation_value(source: str, value: object) -> object:
    if source in INT_FORMAT_FIELDS:
        return _format_int_like(value)
    if source in DECISION_FIELDS:
        return DECISION_REASON_MAP.get(str(value), value)
    if source == "intercept_reason":
        return INTERCEPT_REASON_MAP.get(str(value), value)
    if source == "sku_code_check":
        return SKU_CHECK_MAP.get(str(value), value)
    if source in WARNING_FIELDS:
        return "是" if str(value) == "yes" else "否"
    return value


def _localize_quality_value(source: str, value: object) -> object:
    if source == "type":
        return QUALITY_TYPE_MAP.get(str(value), value)
    if source == "message":
        return QUALITY_MESSAGE_MAP.get(str(value), value)
    return value
