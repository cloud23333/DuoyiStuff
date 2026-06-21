from __future__ import annotations

import io
from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .schema import (
    ACTION_FILL,
    ACTION_STRONG_FILL,
    COLUMN_GROUPS,
    DECIMAL_FORMAT,
    DECIMAL_FORMAT_FIELDS,
    DEFAULT_ROW_HEIGHT_POINTS,
    FORMULA_COLUMN_NAME,
    FORMULA_COLUMN_WIDTH_CHARS,
    GRID_BORDER,
    HEADER_COMMENTS,
    HEADER_FILL_BY_SOURCE,
    HEADER_FONT,
    HEADER_GROUP_FILLS,
    HEADER_ROW_HEIGHT_POINTS,
    INFO_FILL,
    INT_FORMAT_FIELDS,
    INTEGER_FORMAT,
    MAX_COLUMN_WIDTH_CHARS,
    MIN_COLUMN_WIDTH_CHARS,
    MUTED_FILL,
    MUTED_FONT,
    PERCENT_FORMAT,
    PERCENT_FORMAT_FIELDS,
    PLOT_COLUMN_NAME,
    PLOT_COLUMN_WIDTH_CHARS,
    PLOT_IMAGE_HEIGHT_PX,
    PLOT_IMAGE_WIDTH_PX,
    PLOT_ROW_HEIGHT_POINTS,
    RECOMMENDATION_FIELDS,
    RECOMMENDATION_FREEZE_PANES,
    TEXT_WRAP_COLUMN_WIDTH_CHARS,
    TEXT_WRAP_COLUMNS,
    WARNING_FILL,
)


def _write_recommendations_xlsx(
    *,
    path: Path,
    rows: list[dict[str, object]],
    recommendation_rows: list[dict[str, object]],
    plot_cache: dict[tuple[str, str], bytes],
) -> None:
    fields = RECOMMENDATION_FIELDS
    columns = [target for _, target in fields]
    plot_col_idx = columns.index(PLOT_COLUMN_NAME) + 1
    plot_col_letter = get_column_letter(plot_col_idx)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "发货建议明细"

    for col_idx, name in enumerate(columns, start=1):
        worksheet.cell(row=1, column=col_idx, value=name)

    worksheet.column_dimensions[plot_col_letter].width = PLOT_COLUMN_WIDTH_CHARS

    for row_offset, (localized, source) in enumerate(
        zip(rows, recommendation_rows, strict=True)
    ):
        excel_row = row_offset + 2
        for col_idx, name in enumerate(columns, start=1):
            if name == PLOT_COLUMN_NAME:
                continue
            worksheet.cell(row=excel_row, column=col_idx, value=localized.get(name, ""))

        key = (str(source.get("店铺款式编码", "")), str(source.get("店铺商品编码", "")))
        image_bytes = plot_cache.get(key)
        if image_bytes is not None:
            image_stream = io.BytesIO(image_bytes)
            excel_image = XlsxImage(image_stream)
            excel_image.width = PLOT_IMAGE_WIDTH_PX
            excel_image.height = PLOT_IMAGE_HEIGHT_PX
            excel_image.anchor = f"{plot_col_letter}{excel_row}"
            worksheet.add_image(excel_image)
            worksheet.row_dimensions[excel_row].height = PLOT_ROW_HEIGHT_POINTS

    _style_recommendations_worksheet(worksheet=worksheet, fields=fields)
    workbook.save(path)


def _style_recommendations_worksheet(
    *,
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    max_row = worksheet.max_row
    max_col = len(fields)
    last_col_letter = get_column_letter(max_col)
    worksheet.freeze_panes = RECOMMENDATION_FREEZE_PANES
    worksheet.auto_filter.ref = f"A1:{last_col_letter}{max_row}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT_POINTS
    worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT_POINTS

    for col_idx, (source, name) in enumerate(fields, start=1):
        header = worksheet.cell(row=1, column=col_idx)
        header.fill = _header_fill_for_source(source)
        header.font = HEADER_FONT
        header.border = GRID_BORDER
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if name in HEADER_COMMENTS:
            header.comment = Comment(HEADER_COMMENTS[name], "shipment-planner")

    for row_idx in range(2, max_row + 1):
        for col_idx, (source, name) in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = _cell_alignment(source, name)
            cell.border = GRID_BORDER
            if source in INT_FORMAT_FIELDS:
                cell.number_format = INTEGER_FORMAT
            elif source in PERCENT_FORMAT_FIELDS:
                cell.number_format = PERCENT_FORMAT
            elif source in DECIMAL_FORMAT_FIELDS:
                cell.number_format = DECIMAL_FORMAT

    _apply_column_widths(worksheet, fields)
    _apply_column_grouping(worksheet)
    _apply_recommendation_conditional_formatting(worksheet, fields)


def _header_fill_for_source(source: str):
    return HEADER_GROUP_FILLS[HEADER_FILL_BY_SOURCE.get(source, "diagnostic")]


def _cell_alignment(source: str, name: str) -> Alignment:
    return Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=name in TEXT_WRAP_COLUMNS,
    )


def _apply_column_widths(
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    for col_idx, (source, name) in enumerate(fields, start=1):
        letter = get_column_letter(col_idx)
        if name == PLOT_COLUMN_NAME:
            worksheet.column_dimensions[letter].width = PLOT_COLUMN_WIDTH_CHARS
            continue
        width = _recommended_column_width(worksheet, col_idx, source, name)
        worksheet.column_dimensions[letter].width = width


def _apply_column_grouping(worksheet) -> None:
    """Collapse diagnostic groups behind native Excel column outlines."""
    worksheet.sheet_properties.outlinePr.summaryRight = True
    col_idx = 1
    for group in COLUMN_GROUPS:
        start = col_idx
        end = col_idx + len(group.fields) - 1
        if group.outline_level >= 1:
            for col in range(start, end + 1):
                dim = worksheet.column_dimensions[get_column_letter(col)]
                dim.outlineLevel = group.outline_level
                dim.hidden = True
        col_idx = end + 1


def _recommended_column_width(
    worksheet,
    col_idx: int,
    source: str,
    name: str,
) -> float:
    if name == FORMULA_COLUMN_NAME:
        return FORMULA_COLUMN_WIDTH_CHARS
    if name in TEXT_WRAP_COLUMNS:
        return 24.0 if source == "__item_identity__" else TEXT_WRAP_COLUMN_WIDTH_CHARS
    if source == "internal_order_id":
        return 13.0
    if source in {"店铺款式编码", "店铺商品编码"}:
        return 15.0
    if source in {"原始商品编码", "系统商品编码"}:
        return 16.0
    if source in {"decision_reason", "order_decision_reason"}:
        return 8.0
    if source == "line_order_qty":
        return 10.0
    if source == "recommended_ship":
        return 11.0

    max_width = _display_width(name)
    for row_idx in range(2, worksheet.max_row + 1):
        max_width = max(
            max_width,
            _display_width(worksheet.cell(row=row_idx, column=col_idx).value),
        )
    return min(max(max_width + 2, MIN_COLUMN_WIDTH_CHARS), MAX_COLUMN_WIDTH_CHARS)


def _display_width(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    return sum(2 if ord(char) > 127 else 1 for char in text)


def _apply_recommendation_conditional_formatting(
    worksheet,
    fields: Sequence[tuple[str, str]],
) -> None:
    if worksheet.max_row < 2:
        return

    columns = {name: get_column_letter(idx) for idx, (_, name) in enumerate(fields, start=1)}
    last_row = worksheet.max_row

    recommended_col = columns["建议发货量"]
    worksheet.conditional_formatting.add(
        f"{recommended_col}2:{recommended_col}{last_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=ACTION_FILL,
            font=Font(bold=True, color="9C6500"),
        ),
    )
    worksheet.conditional_formatting.add(
        f"{recommended_col}2:{recommended_col}{last_row}",
        CellIsRule(
            operator="equal",
            formula=["0"],
            fill=MUTED_FILL,
            font=MUTED_FONT,
        ),
    )

    notes_col = columns["处理提示"]
    worksheet.conditional_formatting.add(
        f"{notes_col}2:{notes_col}{last_row}",
        FormulaRule(formula=[f"LEN(TRIM(${notes_col}2))>0"], fill=WARNING_FILL),
    )

    for name in (
        "30%内免改数量提示",
        "订单低于起发量提示",
        "小于10不发豁免资格提示",
        "小于10不发豁免生效提示",
        "保底是否触发",
        "订单拦截导致不发提示",
    ):
        col = columns[name]
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            FormulaRule(formula=[f'${col}2="是"'], fill=INFO_FILL),
        )

    check_col = columns["SKU编码校验"]
    worksheet.conditional_formatting.add(
        f"{check_col}2:{check_col}{last_row}",
        FormulaRule(
            formula=[f'OR(${check_col}2="不一致",${check_col}2="缺少销售匹配")'],
            fill=WARNING_FILL,
        ),
    )

    intercept_col = columns["拦截原因"]
    worksheet.conditional_formatting.add(
        f"{intercept_col}2:{intercept_col}{last_row}",
        FormulaRule(formula=[f"LEN(TRIM(${intercept_col}2))>0"], fill=ACTION_STRONG_FILL),
    )
