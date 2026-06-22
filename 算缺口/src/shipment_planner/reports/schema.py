from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Border, Font, PatternFill, Side


@dataclass(frozen=True)
class ColumnGroup:
    """A contiguous block of recommendation columns sharing one header color.

    ``outline_level`` 0 keeps the columns always visible; level 1 collapses
    them behind a native Excel outline (expand on demand).
    """

    key: str
    fill: str
    outline_level: int
    title: str
    fields: tuple[tuple[str, str], ...]


COLUMN_GROUPS: tuple[ColumnGroup, ...] = (
    ColumnGroup(
        "core",
        "action",
        0,
        "核心决策",
        (
            ("internal_order_id", "订单号"),
            ("__item_identity__", "商品定位"),
            ("line_order_qty", "订单行数量"),
            ("recommended_ship", "建议发货量"),
            ("decision_reason", "SKU建议"),
            ("order_decision_reason", "订单建议"),
            ("__action_notes__", "处理提示"),
            ("gap", "缺口"),
            ("__available_stock__", "可用库存"),
        ),
    ),
    ColumnGroup(
        "stock",
        "stock",
        1,
        "库存明细",
        (
            ("key_order_qty", "同款同SKU总下单量"),
            ("key_recommended_total", "同款同SKU建议总量"),
            ("wh", "平台仓内库存"),
            ("pending_recv", "平台待收货库存"),
            ("shipping_in_progress", "发货中数量"),
            ("pending_ship", "平台待发货库存"),
        ),
    ),
    ColumnGroup(
        "forecast",
        "forecast",
        1,
        "预测明细",
        (
            ("forecast_stocking_period_sales", "预测备货期销量"),
            ("forecast_daily_sales", "预测日均销量"),
            ("effective_daily_sales", "异常调整后日均销量"),
            ("stocking_days", "备货逻辑天数"),
            ("forecast_strategy", "预测策略"),
            ("demand_profile", "需求类型"),
            ("anomaly_flags", "异常标记"),
            ("service_level", "服务水平"),
            ("forecast_model", "预测模型"),
            ("sku_forecast_abs_error", "SKU预测绝对误差"),
            ("sku_forecast_signed_error", "SKU预测偏差"),
        ),
    ),
    ColumnGroup(
        "rule",
        "rule",
        1,
        "规则明细",
        (
            ("recommended_ship_before_small_change_rule", "30%规则前建议发货量"),
            ("small_change_ratio_before_rule", "30%规则前变动比例"),
            ("small_change_keep_warning", "30%内免改数量提示"),
            ("order_recommended_ship_total_before_threshold", "订单阈值前建议总量"),
            ("min_order_ship_qty_threshold", "最小发货阈值"),
            ("order_low_qty_warning", "订单低于起发量提示"),
            ("min_order_ship_qty_exempt_warning", "小于10不发豁免资格提示"),
            ("min_order_ship_qty_exempt_applied_warning", "小于10不发豁免生效提示"),
            ("base_stock_qty", "保底库存目标"),
            ("base_stock_gap", "保底库存缺口"),
            ("base_stock_triggered_warning", "保底是否触发"),
            ("intercept_reason", "拦截原因"),
            ("order_intercept_warning", "订单拦截导致不发提示"),
        ),
    ),
    ColumnGroup(
        "identity",
        "diagnostic",
        1,
        "编码校验",
        (
            ("店铺款式编码", "店铺款式编码"),
            ("店铺商品编码", "店铺商品编码"),
            ("原始商品编码", "订单商品编码"),
            ("系统商品编码", "系统商品编码"),
            ("sku_code_check", "SKU编码校验"),
        ),
    ),
    ColumnGroup(
        "explain",
        "action",
        0,
        "推导与曲线",
        (
            ("__plot__", "销量与预测曲线"),
            ("__formula__", "缺口与发货量推导"),
        ),
    ),
)

RECOMMENDATION_FIELDS: list[tuple[str, str]] = [
    field for group in COLUMN_GROUPS for field in group.fields
]
HEADER_FILL_BY_SOURCE: dict[str, str] = {
    source: group.fill for group in COLUMN_GROUPS for source, _ in group.fields
}

FORMULA_COLUMN_NAME = "缺口与发货量推导"
FORMULA_COLUMN_WIDTH_CHARS = 48.0

PLOT_COLUMN_NAME = "销量与预测曲线"
PLOT_COLUMN_WIDTH_CHARS = 46.0
PLOT_ROW_HEIGHT_POINTS = 100.0
PLOT_IMAGE_WIDTH_PX = 320
PLOT_IMAGE_HEIGHT_PX = 128

RECOMMENDATION_FREEZE_PANES = "E2"
HEADER_ROW_HEIGHT_POINTS = 24.0
DEFAULT_ROW_HEIGHT_POINTS = 34.0
MIN_COLUMN_WIDTH_CHARS = 8.0
MAX_COLUMN_WIDTH_CHARS = 34.0
TEXT_WRAP_COLUMN_WIDTH_CHARS = 30.0

HEADER_COMMENTS = {
    "建议发货量": "最终建议执行数量，优先查看或复制此列。",
    "商品定位": "合并显示店铺款式编码、店铺商品编码和订单商品编码，减少冻结列占用。",
    "处理提示": "汇总影响建议量的主要原因和异常提示。",
    "可用库存": "平台仓内库存 + 平台待收货库存 + 发货中数量。",
    "缺口": "预测备货期需求扣减可用库存后的缺口。",
}

INTEGER_FORMAT = "#,##0"
DECIMAL_FORMAT = "#,##0.####"
PERCENT_FORMAT = "0%"

PERCENT_FORMAT_FIELDS = {
    "service_level",
    "small_change_ratio_before_rule",
}

DECIMAL_FORMAT_FIELDS = {
    "effective_daily_sales",
    "forecast_daily_sales",
    "forecast_stocking_period_sales",
    "stocking_days",
    "sku_forecast_abs_error",
    "sku_forecast_signed_error",
    "wh",
    "pending_recv",
    "shipping_in_progress",
    "pending_ship",
    "__available_stock__",
}

TEXT_WRAP_COLUMNS = {
    "商品定位",
    "处理提示",
    "异常标记",
    "缺口与发货量推导",
}

HEADER_GROUP_FILLS = {
    "action": PatternFill("solid", fgColor="1F4E78"),
    "stock": PatternFill("solid", fgColor="5B7F55"),
    "rule": PatternFill("solid", fgColor="8064A2"),
    "forecast": PatternFill("solid", fgColor="4F81BD"),
    "diagnostic": PatternFill("solid", fgColor="7F6000"),
}
HEADER_FONT = Font(color="FFFFFF", bold=True)
GRID_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ACTION_FILL = PatternFill("solid", fgColor="FFF2CC")
ACTION_STRONG_FILL = PatternFill("solid", fgColor="FCE4D6")
MUTED_FILL = PatternFill("solid", fgColor="F2F2F2")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
INFO_FILL = PatternFill("solid", fgColor="E2F0D9")
MUTED_FONT = Font(color="808080")

QUALITY_FIELDS = [
    ("type", "问题类型"),
    ("internal_order_id", "内部订单号"),
    ("skc", "店铺款式编码"),
    ("skuid", "店铺商品编码"),
    ("order_sku", "原始商品编码"),
    ("system_sku", "系统商品编码"),
    ("message", "问题说明"),
]

DECISION_REASON_MAP = {
    "ship_all": "全发",
    "ship_partial": "部分发",
    "hold": "暂不发",
}

INTERCEPT_REASON_MAP = {
    "skc": "命中SKC拦截",
    "skuid": "命中SKUID拦截",
    "skc_and_skuid": "命中SKC和SKUID拦截",
}

SKU_CHECK_MAP = {
    "exact_match": "完全一致",
    "normalized_match": "标准化一致",
    "diff": "不一致",
    "missing_key": "缺少销售匹配",
}

QUALITY_TYPE_MAP = {
    "sku_code_diff": "SKU编码不一致",
    "missing_sales_key": "缺少销售匹配",
    "missing_daily_sales": "缺少每日销量数据",
}

QUALITY_MESSAGE_MAP = {
    "Order SKU and system SKU differ after normalization": "原始商品编码与系统商品编码在标准化后仍不一致",
    "No sales row found for (SKC, SKUID)": "未找到对应销售键 (SKC, SKUID)",
    "Missing daily sales data for (SKC, SKUID)": "未找到对应每日销量数据 (SKC, SKUID)",
}

INT_FORMAT_FIELDS = {
    "line_order_qty",
    "key_order_qty",
    "gap",
    "base_stock_qty",
    "base_stock_gap",
    "key_recommended_total",
    "recommended_ship_before_small_change_rule",
    "recommended_ship",
    "order_recommended_ship_total_before_threshold",
    "min_order_ship_qty_threshold",
}

SUMMARY_INT_FORMAT_FIELDS = {
    "order_lines",
    "sales_rows",
    "matched_order_lines",
    "total_order_qty",
    "total_recommended_qty",
    "small_change_kept_lines",
    "base_stock_qty",
    "base_stock_triggered_skus",
    "base_stock_triggered_lines",
    "quality_issue_rows",
    "duplicate_sales_keys",
    "min_order_ship_qty_threshold",
    "low_qty_orders_before_exempt",
    "low_qty_order_lines_before_exempt",
    "low_qty_orders",
    "low_qty_order_lines",
    "low_qty_orders_exempted",
    "low_qty_order_lines_exempted",
    "sku_order_limit_rule_count",
    "sku_order_limit_capped_lines",
    "excluded_skc_rule_count",
    "excluded_skuid_rule_count",
    "intercepted_order_lines",
    "intercepted_orders",
}

DECISION_FIELDS = {"decision_reason", "order_decision_reason"}

WARNING_FIELDS = {
    "order_low_qty_warning",
    "min_order_ship_qty_exempt_warning",
    "min_order_ship_qty_exempt_applied_warning",
    "order_intercept_warning",
    "small_change_keep_warning",
    "base_stock_triggered_warning",
}

SUMMARY_FIELDS = [
    ("order_lines", "订单行数"),
    ("sales_rows", "销售行数"),
    ("matched_order_lines", "匹配订单行数"),
    ("join_coverage_pct", "匹配覆盖率_百分比"),
    ("total_order_qty", "总下单量"),
    ("total_recommended_qty", "建议发货总量"),
    ("small_change_kept_lines", "触发30%免改行数"),
    ("decision_ship_all", "建议_全发_行数"),
    ("decision_ship_partial", "建议_部分发_行数"),
    ("decision_hold", "建议_暂不发_行数"),
    ("sku_check_exact_match", "SKU校验_完全一致_行数"),
    ("sku_check_normalized_match", "SKU校验_标准化一致_行数"),
    ("sku_check_diff", "SKU校验_不一致_行数"),
    ("sku_check_missing_key", "SKU校验_缺少销售匹配_行数"),
    ("quality_issue_rows", "质量问题行数"),
    ("duplicate_sales_keys", "销售重复键数量"),
    ("demand_profile_summary", "需求类型分布"),
    ("anomaly_flag_summary", "异常标记分布"),
    ("service_level_summary", "服务水平分布"),
    ("forecast_model_summary", "预测模型分布"),
    ("service_level_offset", "全局服务水平偏移"),
    ("base_stock_qty", "保底库存目标"),
    ("base_stock_triggered_skus", "保底触发SKU数"),
    ("base_stock_triggered_lines", "保底触发订单行数"),
    ("min_order_ship_qty_threshold", "最小发货阈值"),
    ("low_qty_orders_before_exempt", "阈值前低发货量订单数"),
    ("low_qty_order_lines_before_exempt", "阈值前低发货量订单行数"),
    ("low_qty_orders", "低于阈值订单数_提示"),
    ("low_qty_order_lines", "低于阈值订单行数_提示"),
    ("low_qty_orders_exempted", "低于阈值豁免订单数"),
    ("low_qty_order_lines_exempted", "低于阈值豁免订单行数"),
    ("sku_order_limit_rule_count", "订单内SKU限额规则数"),
    ("sku_order_limit_capped_lines", "触发订单内SKU限额行数"),
    ("excluded_skc_rule_count", "SKC拦截规则数"),
    ("excluded_skuid_rule_count", "SKUID拦截规则数"),
    ("intercepted_order_lines", "命中拦截订单行数"),
    ("intercepted_orders", "拦截导致不发订单数"),
]

FORECAST_EVAL_SUMMARY_KEY = "预测回测评估"
FORECAST_EVAL_SUMMARY_FIELDS = [
    ("evaluated_skus", "评估SKU数"),
    ("skipped_insufficient_history", "历史不足跳过SKU数"),
    ("holdout_days", "留出天数"),
    ("mae", "MAE"),
    ("wape", "WAPE"),
    ("bias", "平均偏差"),
]
FORECAST_EVAL_STRATEGY_FIELDS = [
    ("count", "SKU数"),
    ("mae", "MAE"),
    ("wape", "WAPE"),
    ("bias", "平均偏差"),
]
FORECAST_EVAL_INT_FIELDS = {
    "evaluated_skus",
    "skipped_insufficient_history",
    "holdout_days",
    "count",
}
