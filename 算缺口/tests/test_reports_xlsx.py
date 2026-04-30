from __future__ import annotations

import sys
import types

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from shipment_planner.reports import (
    _build_plot_cache,
    _localize_recommendation_row,
    _write_recommendations_xlsx,
)


def test_recommendation_xlsx_is_execution_first(tmp_path):
    path = _write_test_recommendation_report(tmp_path)

    workbook = load_workbook(path)
    worksheet = workbook["发货建议明细"]
    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]

    assert headers[:10] == [
        "订单号",
        "商品定位",
        "订单行数量",
        "建议发货量",
        "SKU建议",
        "订单建议",
        "处理提示",
        "店铺款式编码",
        "店铺商品编码",
        "订单商品编码",
    ]
    assert headers.index("建议发货量") == 3
    assert worksheet.freeze_panes == "E2"
    assert worksheet.auto_filter.ref == f"A1:{get_column_letter(worksheet.max_column)}2"

    values = {
        header: worksheet.cell(row=2, column=idx + 1).value
        for idx, header in enumerate(headers)
    }
    assert values["建议发货量"] == 6
    assert values["商品定位"] == "SKC skc-1\nSKUID skuid-1\nSKU SKU-A"
    assert values["SKU建议"] == "部分发"
    assert values["订单建议"] == "部分发"
    assert values["可用库存"] == 6
    assert values["SKU编码校验"] == "不一致"
    assert values["处理提示"] == (
        "SKU不一致；订单低于起发量；小于10豁免生效；30%内免改；保底触发；销量异常：孤立爆单"
    )


def test_recommendation_xlsx_has_readability_formatting(tmp_path):
    path = _write_test_recommendation_report(tmp_path)

    workbook = load_workbook(path)
    worksheet = workbook["发货建议明细"]
    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    columns = {header: idx + 1 for idx, header in enumerate(headers)}

    assert worksheet.cell(row=2, column=columns["建议发货量"]).number_format == "#,##0"
    assert worksheet.cell(row=2, column=columns["订单行数量"]).alignment.horizontal == "center"
    assert worksheet.cell(row=2, column=columns["商品定位"]).alignment.horizontal == "center"
    assert worksheet.cell(row=2, column=columns["服务水平"]).number_format == "0%"
    assert worksheet.cell(row=2, column=columns["30%规则前变动比例"]).number_format == "0%"
    assert worksheet.cell(row=1, column=columns["建议发货量"]).comment is not None
    assert worksheet.cell(row=1, column=columns["商品定位"]).comment is not None
    assert worksheet.cell(row=1, column=columns["处理提示"]).comment is not None
    assert worksheet.cell(row=2, column=columns["建议发货量"]).border.left.style == "thin"
    assert worksheet.column_dimensions[get_column_letter(columns["销量与预测曲线"])].width == 46.0
    assert len(worksheet.conditional_formatting) > 0


def test_recommendation_plot_title_includes_forecast_model(monkeypatch):
    captured: dict[str, object] = {}

    def fake_render_sku_plot(**kwargs):
        captured.update(kwargs)
        return b"png"

    fake_plots = types.ModuleType("shipment_planner.plots")
    fake_plots.render_sku_plot = fake_render_sku_plot
    monkeypatch.setitem(sys.modules, "shipment_planner.plots", fake_plots)

    _build_plot_cache(
        [_recommendation_row()],
        {("skc-1", "skuid-1"): (1, 2, 3)},
    )

    assert captured["title"] == "SKC:skc-1 / SKUID:skuid-1 | tsb"
    assert captured["forecast_daily_sales"] == 3.0


def _write_test_recommendation_report(tmp_path):
    source_rows = [_recommendation_row()]
    localized_rows = [_localize_recommendation_row(row) for row in source_rows]
    path = tmp_path / "发货建议明细.xlsx"
    _write_recommendations_xlsx(
        path=path,
        rows=localized_rows,
        recommendation_rows=source_rows,
        plot_cache={},
    )
    return path


def _recommendation_row() -> dict[str, object]:
    return {
        "internal_order_id": "order-1",
        "店铺款式编码": "skc-1",
        "店铺商品编码": "skuid-1",
        "原始商品编码": "SKU-A",
        "系统商品编码": "SKU-B",
        "sku_code_check": "diff",
        "line_order_qty": 10,
        "key_order_qty": 20,
        "forecast_strategy": "激进",
        "demand_profile": "波动款",
        "anomaly_flags": "孤立爆单",
        "service_level": 0.7,
        "forecast_model": "tsb",
        "effective_daily_sales": 2.5,
        "forecast_daily_sales": 3.0,
        "forecast_stocking_period_sales": 21.0,
        "stocking_days": 7,
        "wh": 1,
        "pending_recv": 2,
        "shipping_in_progress": 3,
        "pending_ship": 4,
        "gap": 15,
        "base_stock_qty": 2,
        "base_stock_gap": 1,
        "base_stock_triggered_warning": "yes",
        "key_recommended_total": 6,
        "recommended_ship_before_small_change_rule": 4,
        "small_change_ratio_before_rule": 0.4,
        "small_change_keep_warning": "yes",
        "recommended_ship": 6,
        "decision_reason": "ship_partial",
        "order_decision_reason": "ship_partial",
        "intercept_reason": "",
        "order_intercept_warning": "no",
        "order_recommended_ship_total_before_threshold": 6,
        "min_order_ship_qty_threshold": 10,
        "order_low_qty_warning": "yes",
        "min_order_ship_qty_exempt_warning": "yes",
        "min_order_ship_qty_exempt_applied_warning": "yes",
    }
