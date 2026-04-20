from __future__ import annotations

import json

from openpyxl import load_workbook

from shipment_planner.eval_forecast import EvalRow
from shipment_planner import reports


def _recommendation_row(
    *,
    skc: str = "skc-1",
    skuid: str = "sku-1",
    decision_reason: str = "hold",
    gap: int = 0,
) -> dict[str, object]:
    return {
        "internal_order_id": "order-1",
        "店铺款式编码": skc,
        "店铺商品编码": skuid,
        "原始商品编码": "SKU-1",
        "系统商品编码": "SKU-1",
        "sku_code_check": "exact_match",
        "line_order_qty": 10,
        "key_order_qty": 10,
        "forecast_strategy": "正常",
        "forecast_daily_sales": 2.0,
        "forecast_stocking_period_sales": 20.0,
        "stocking_days": 10,
        "wh": 20,
        "pending_ship": 0,
        "shipping_in_progress": 0,
        "pending_recv": 0,
        "gap": gap,
        "key_recommended_total": 0,
        "recommended_ship_before_small_change_rule": 0,
        "small_change_ratio_before_rule": 0,
        "small_change_keep_warning": "no",
        "recommended_ship": 0,
        "decision_reason": decision_reason,
        "order_decision_reason": decision_reason,
        "intercept_reason": "",
        "order_intercept_warning": "no",
        "order_recommended_ship_total_before_threshold": 0,
        "min_order_ship_qty_threshold": 10,
        "order_low_qty_warning": "no",
        "min_order_ship_qty_exempt_warning": "no",
        "min_order_ship_qty_exempt_applied_warning": "no",
    }


def _summary() -> dict[str, object]:
    return {
        "order_lines": 1,
        "sales_rows": 1,
        "matched_order_lines": 1,
        "join_coverage_pct": 100.0,
        "total_order_qty": 10,
        "total_recommended_qty": 0,
        "small_change_kept_lines": 0,
        "decision_ship_all": 0,
        "decision_ship_partial": 0,
        "decision_hold": 1,
        "sku_check_exact_match": 1,
        "sku_check_normalized_match": 0,
        "sku_check_diff": 0,
        "sku_check_missing_key": 0,
        "quality_issue_rows": 0,
        "duplicate_sales_keys": 0,
        "global_gap_multiplier": 1.0,
        "min_order_ship_qty_threshold": 10,
        "low_qty_orders_before_exempt": 0,
        "low_qty_order_lines_before_exempt": 0,
        "low_qty_orders": 0,
        "low_qty_order_lines": 0,
        "low_qty_orders_exempted": 0,
        "low_qty_order_lines_exempted": 0,
        "sku_order_limit_rule_count": 0,
        "sku_order_limit_capped_lines": 0,
        "excluded_skc_rule_count": 0,
        "excluded_skuid_rule_count": 0,
        "intercepted_order_lines": 0,
        "intercepted_orders": 0,
    }


def test_export_reports_removes_stale_recommendation_csv(tmp_path) -> None:
    stale_csv = tmp_path / "发货建议明细.csv"
    stale_csv.write_text("stale,data\nold,value\n", encoding="utf-8")

    outputs = reports.export_reports(
        tmp_path,
        [_recommendation_row()],
        [],
        _summary(),
        daily_sales_by_key={},
    )

    assert "final_recommendation_csv" not in outputs
    assert not stale_csv.exists()


def test_export_reports_merges_forecast_eval_into_run_summary(tmp_path) -> None:
    reports.export_reports(
        tmp_path,
        [_recommendation_row()],
        [],
        _summary(),
        daily_sales_by_key={},
        eval_summary={
            "evaluated_skus": 3,
            "skipped_insufficient_history": 2,
            "holdout_days": 7,
            "mae": 1.25,
            "wape": 0.5,
            "bias": -0.25,
            "by_strategy": {
                "正常": {"count": 3, "mae": 1.25, "wape": 0.5, "bias": -0.25}
            },
        },
    )

    summary = json.loads((tmp_path / "运行摘要.json").read_text(encoding="utf-8"))

    assert summary["预测回测评估"]["评估SKU数"] == 3
    assert summary["预测回测评估"]["历史不足跳过SKU数"] == 2
    assert summary["预测回测评估"]["按策略"]["正常"]["SKU数"] == 3


def test_export_reports_attaches_per_sku_forecast_error_to_rows(tmp_path) -> None:
    reports.export_reports(
        tmp_path,
        [_recommendation_row()],
        [],
        _summary(),
        daily_sales_by_key={},
        eval_rows=[
            EvalRow(
                skc="skc-1",
                skuid="sku-1",
                strategy="正常",
                train_days=14,
                holdout_days=7,
                forecast_daily_sales=2.0,
                forecast_holdout_total=14.0,
                actual_holdout_total=10,
                abs_error=4.0,
                signed_error=4.0,
            )
        ],
    )

    workbook = load_workbook(tmp_path / "发货建议明细.xlsx", read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    detail = dict(zip(header, rows[1]))

    assert detail["SKU预测绝对误差"] == 4.0
    assert detail["SKU预测偏差"] == 4.0


def test_build_plot_cache_renders_every_sku_row(monkeypatch) -> None:
    rendered_titles: list[str] = []

    def fake_render_sku_plot(**kwargs) -> bytes:
        rendered_titles.append(str(kwargs["title"]))
        return b"plot-bytes"

    monkeypatch.setattr(reports, "render_sku_plot", fake_render_sku_plot)
    rows = [
        _recommendation_row(
            skc=f"skc-{idx:03d}",
            skuid=f"sku-{idx:03d}",
            decision_reason="hold" if idx % 2 else "ship_partial",
            gap=0 if idx % 2 else idx + 1,
        )
        for idx in range(70)
    ]
    daily_sales_by_key = {
        (str(row["店铺款式编码"]), str(row["店铺商品编码"])): (1, 2, 3)
        for row in rows
    }

    cache = reports._build_plot_cache(rows, daily_sales_by_key)

    assert len(cache) == len(rows)
    assert len(rendered_titles) == len(rows)
    assert "SKC:skc-000 / SKUID:sku-000" in rendered_titles
    assert (
        f"SKC:skc-{len(rows) - 1:03d} / SKUID:sku-{len(rows) - 1:03d}"
        in rendered_titles
    )
