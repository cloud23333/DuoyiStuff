"""One-off diff between the old (CSV) and new (XLSX) 发货建议明细 outputs.

Usage::

    python scripts/compare_outputs.py \
        --old data/output/发货建议明细.csv \
        --new data/output/发货建议明细.xlsx \
        --out data/output/算法对比.xlsx

Matches rows on (内部订单号, 店铺款式编码, 店铺商品编码) + occurrence index, then
reports per-row changes for the numeric/strategy columns that our engine edits.
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

KEY_COLUMNS = ["内部订单号", "店铺款式编码", "店铺商品编码"]
LABEL_COLUMNS = [
    "内部订单号",
    "店铺款式编码",
    "店铺商品编码",
    "原始商品编码",
    "订单行数量",
]
STRING_DIFF_COLUMNS = ["预测策略", "SKU建议类型"]
NUMERIC_DIFF_COLUMNS = [
    "预测日均销量",
    "预测备货期销量",
    "缺口",
    "建议发货量",
]

COMPARE_COLUMNS = STRING_DIFF_COLUMNS + NUMERIC_DIFF_COLUMNS

DIFF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FONT = Font(bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--old", required=True, help="Old CSV path")
    parser.add_argument("--new", required=True, help="New XLSX path")
    parser.add_argument("--out", required=True, help="Output XLSX path for the diff")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    old_rows = _read_csv(Path(args.old))
    new_rows = _read_xlsx(Path(args.new))

    old_indexed = _index_rows(old_rows)
    new_indexed = _index_rows(new_rows)

    all_keys = list(dict.fromkeys(list(old_indexed.keys()) + list(new_indexed.keys())))

    diff_rows: list[dict[str, Any]] = []
    summary: dict[str, Any] = {
        "total_keys": len(all_keys),
        "only_in_old": 0,
        "only_in_new": 0,
        "in_both": 0,
        "rows_with_any_change": 0,
        "per_column_changed": Counter(),
        "strategy_transitions": Counter(),
        "decision_transitions": Counter(),
        "total_recommended_old": 0.0,
        "total_recommended_new": 0.0,
        "total_gap_old": 0.0,
        "total_gap_new": 0.0,
    }

    for key in all_keys:
        old_row = old_indexed.get(key)
        new_row = new_indexed.get(key)

        if old_row is None:
            summary["only_in_new"] += 1
        elif new_row is None:
            summary["only_in_old"] += 1
        else:
            summary["in_both"] += 1

        diff = _build_diff_row(key=key, old_row=old_row, new_row=new_row)
        if _row_has_any_change(diff):
            summary["rows_with_any_change"] += 1
            for column in COMPARE_COLUMNS:
                if diff.get(f"{column}__changed"):
                    summary["per_column_changed"][column] += 1
            _track_transitions(diff, summary)

        summary["total_recommended_old"] += _as_float(
            (old_row or {}).get("建议发货量", 0)
        )
        summary["total_recommended_new"] += _as_float(
            (new_row or {}).get("建议发货量", 0)
        )
        summary["total_gap_old"] += _as_float((old_row or {}).get("缺口", 0))
        summary["total_gap_new"] += _as_float((new_row or {}).get("缺口", 0))

        diff_rows.append(diff)

    _write_report(
        out_path=Path(args.out),
        diff_rows=diff_rows,
        summary=summary,
    )

    print(f"Wrote diff report: {args.out}")
    _print_summary(summary)
    return 0


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        return []

    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        if all(value is None or value == "" for value in values):
            continue
        row: dict[str, Any] = {}
        for column_name, value in zip(header, values):
            if column_name == "销量与预测曲线":
                continue
            row[column_name] = "" if value is None else value
        rows.append(row)
    return rows


def _index_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str, int], dict[str, Any]]:
    occurrence: dict[tuple[str, str, str], int] = defaultdict(int)
    indexed: dict[tuple[str, str, str, int], dict[str, Any]] = {}
    for row in rows:
        base_key = tuple(str(row.get(column, "")) for column in KEY_COLUMNS)
        occurrence_idx = occurrence[base_key]
        occurrence[base_key] += 1
        indexed[(*base_key, occurrence_idx)] = row
    return indexed


def _build_diff_row(
    *,
    key: tuple[str, str, str, int],
    old_row: dict[str, Any] | None,
    new_row: dict[str, Any] | None,
) -> dict[str, Any]:
    internal_order_id, skc, skuid, occurrence_idx = key
    source = new_row if new_row is not None else (old_row or {})

    diff: dict[str, Any] = {
        "内部订单号": internal_order_id,
        "店铺款式编码": skc,
        "店铺商品编码": skuid,
        "出现序号": occurrence_idx,
        "原始商品编码": source.get("原始商品编码", ""),
        "订单行数量": _to_number(source.get("订单行数量", "")),
        "存在": _presence_tag(old_row, new_row),
    }

    for column in COMPARE_COLUMNS:
        old_value = _coerce(old_row.get(column) if old_row is not None else "")
        new_value = _coerce(new_row.get(column) if new_row is not None else "")
        diff[f"{column}__old"] = old_value
        diff[f"{column}__new"] = new_value

        if column in NUMERIC_DIFF_COLUMNS:
            delta = _as_float(new_value) - _as_float(old_value)
            diff[f"{column}__delta"] = round(delta, 4)
            diff[f"{column}__changed"] = not _close(
                _as_float(old_value), _as_float(new_value)
            )
        else:
            diff[f"{column}__changed"] = str(old_value) != str(new_value)

    return diff


def _row_has_any_change(diff: dict[str, Any]) -> bool:
    return any(diff.get(f"{column}__changed") for column in COMPARE_COLUMNS)


def _track_transitions(diff: dict[str, Any], summary: dict[str, Any]) -> None:
    if diff.get("预测策略__changed"):
        summary["strategy_transitions"][
            f"{diff['预测策略__old'] or '(空)'} → {diff['预测策略__new'] or '(空)'}"
        ] += 1
    if diff.get("SKU建议类型__changed"):
        summary["decision_transitions"][
            f"{diff['SKU建议类型__old'] or '(空)'} → {diff['SKU建议类型__new'] or '(空)'}"
        ] += 1


def _presence_tag(
    old_row: dict[str, Any] | None, new_row: dict[str, Any] | None
) -> str:
    if old_row is not None and new_row is not None:
        return "新旧都有"
    if old_row is None:
        return "仅新版"
    return "仅旧版"


def _coerce(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _to_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _close(a: float, b: float, tolerance: float = 1e-6) -> bool:
    return abs(a - b) <= tolerance


def _write_report(
    *,
    out_path: Path,
    diff_rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "逐行对比"
    _write_detail_sheet(detail_sheet, diff_rows)

    summary_sheet = workbook.create_sheet("汇总")
    _write_summary_sheet(summary_sheet, summary)

    workbook.save(out_path)


def _detail_columns() -> list[str]:
    columns = [
        "内部订单号",
        "店铺款式编码",
        "店铺商品编码",
        "出现序号",
        "原始商品编码",
        "订单行数量",
        "存在",
    ]
    for column in COMPARE_COLUMNS:
        columns.append(f"{column}__old")
        columns.append(f"{column}__new")
        if column in NUMERIC_DIFF_COLUMNS:
            columns.append(f"{column}__delta")
    columns.append("任一变化")
    return columns


def _write_detail_sheet(worksheet: Any, diff_rows: Iterable[dict[str, Any]]) -> None:
    columns = _detail_columns()
    for col_idx, name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT

    for row_idx, diff in enumerate(diff_rows, start=2):
        any_change = _row_has_any_change(diff)
        worksheet.cell(row=row_idx, column=len(columns), value="是" if any_change else "否")
        for col_idx, name in enumerate(columns[:-1], start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=diff.get(name, ""))
            if name.endswith("__new") or name.endswith("__delta"):
                column_base = name.rsplit("__", 1)[0]
                if diff.get(f"{column_base}__changed"):
                    cell.fill = DIFF_FILL

    worksheet.freeze_panes = "A2"
    for col_idx, name in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(12, len(name) + 2)


def _write_summary_sheet(worksheet: Any, summary: dict[str, Any]) -> None:
    rows: list[tuple[str, Any]] = [
        ("匹配的行数（新旧都有）", summary["in_both"]),
        ("仅旧版存在", summary["only_in_old"]),
        ("仅新版存在", summary["only_in_new"]),
        ("发生任意变化的行数", summary["rows_with_any_change"]),
        ("旧版建议发货总量", round(summary["total_recommended_old"], 2)),
        ("新版建议发货总量", round(summary["total_recommended_new"], 2)),
        (
            "建议发货总量差（新-旧）",
            round(summary["total_recommended_new"] - summary["total_recommended_old"], 2),
        ),
        ("旧版缺口总量", round(summary["total_gap_old"], 2)),
        ("新版缺口总量", round(summary["total_gap_new"], 2)),
        (
            "缺口总量差（新-旧）",
            round(summary["total_gap_new"] - summary["total_gap_old"], 2),
        ),
    ]
    worksheet.cell(row=1, column=1, value="指标").font = HEADER_FONT
    worksheet.cell(row=1, column=2, value="值").font = HEADER_FONT
    for row_idx, (label, value) in enumerate(rows, start=2):
        worksheet.cell(row=row_idx, column=1, value=label)
        worksheet.cell(row=row_idx, column=2, value=value)

    current_row = len(rows) + 3
    worksheet.cell(row=current_row, column=1, value="按列变动统计").font = HEADER_FONT
    current_row += 1
    for column, count in summary["per_column_changed"].most_common():
        worksheet.cell(row=current_row, column=1, value=column)
        worksheet.cell(row=current_row, column=2, value=count)
        current_row += 1

    current_row += 1
    worksheet.cell(row=current_row, column=1, value="预测策略迁移").font = HEADER_FONT
    current_row += 1
    for transition, count in summary["strategy_transitions"].most_common():
        worksheet.cell(row=current_row, column=1, value=transition)
        worksheet.cell(row=current_row, column=2, value=count)
        current_row += 1

    current_row += 1
    worksheet.cell(row=current_row, column=1, value="SKU建议类型迁移").font = HEADER_FONT
    current_row += 1
    for transition, count in summary["decision_transitions"].most_common():
        worksheet.cell(row=current_row, column=1, value=transition)
        worksheet.cell(row=current_row, column=2, value=count)
        current_row += 1

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 16


def _print_summary(summary: dict[str, Any]) -> None:
    print(f"  新旧都有: {summary['in_both']}")
    print(f"  仅旧版: {summary['only_in_old']}  仅新版: {summary['only_in_new']}")
    print(f"  发生变化的行: {summary['rows_with_any_change']}")
    print(
        "  建议发货总量: 旧 "
        f"{summary['total_recommended_old']:.0f} → 新 "
        f"{summary['total_recommended_new']:.0f} "
        f"(Δ {summary['total_recommended_new'] - summary['total_recommended_old']:+.0f})"
    )
    print(
        "  缺口总量: 旧 "
        f"{summary['total_gap_old']:.0f} → 新 "
        f"{summary['total_gap_new']:.0f} "
        f"(Δ {summary['total_gap_new'] - summary['total_gap_old']:+.0f})"
    )
    if summary["per_column_changed"]:
        print("  按列变化:")
        for column, count in summary["per_column_changed"].most_common():
            print(f"    {column}: {count}")
    if summary["strategy_transitions"]:
        print("  预测策略迁移（前 5）:")
        for transition, count in summary["strategy_transitions"].most_common(5):
            print(f"    {transition}: {count}")


if __name__ == "__main__":
    raise SystemExit(main())
