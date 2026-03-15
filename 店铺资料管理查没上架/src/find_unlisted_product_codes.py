#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel


DEFAULT_TIME_FORMAT = "%Y-%m-%d %H:%M:%S"
TIME_FORMATS = (
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d",
    "%Y-%m-%d",
)
DETAIL_HEADERS = ["店铺名称", "商品编码", "创建时间"]
SUMMARY_HEADERS = ["店铺名称", "未出现商品数"]
CHINESE_CHAR_PATTERN = re.compile(r"[\u4e00-\u9fff]")


@dataclass
class TimeValue:
    raw: str
    parsed: Optional[datetime]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}".rstrip("0").rstrip(".")
    return normalize_text(value)


def should_exclude_code(code: str) -> bool:
    if not code:
        return True
    if code.upper().startswith("FH"):
        return True
    if CHINESE_CHAR_PATTERN.search(code):
        return True
    return False


def parse_datetime(value: object) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt
            if isinstance(dt, date):
                return datetime(dt.year, dt.month, dt.day)
        except Exception:
            pass

    text = normalize_text(value)
    if not text:
        return None
    for fmt in TIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def resolve_column(columns: Iterable[str], target: str) -> str:
    normalized_target = target.strip()
    for col in columns:
        if col.strip() == normalized_target:
            return col
    raise ValueError(f"找不到列: {target}。当前列为: {list(columns)}")


def resolve_column_or_none(columns: Iterable[str], target: str) -> Optional[str]:
    try:
        return resolve_column(columns, target)
    except ValueError:
        return None


def read_excel(path: Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {sheet_name} ({path})")

    ws = wb[sheet_name]
    ws.reset_dimensions()

    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration as e:
        raise ValueError(f"空表: {path}") from e

    columns = [normalize_text(c) for c in header_row]
    if not any(columns):
        raise ValueError(f"表头为空: {path}")

    rows: List[Dict[str, object]] = []
    for raw_row in row_iter:
        if raw_row is None:
            continue
        row_data: Dict[str, object] = {}
        has_value = False
        for i, col in enumerate(columns):
            if not col:
                continue
            cell = raw_row[i] if i < len(raw_row) else None
            row_data[col] = cell
            if normalize_text(cell):
                has_value = True
        if has_value:
            rows.append(row_data)

    return columns, rows


def choose_latest_time(current: TimeValue, candidate: TimeValue) -> TimeValue:
    if current.parsed is None and candidate.parsed is not None:
        return candidate
    if current.parsed is not None and candidate.parsed is not None and candidate.parsed > current.parsed:
        return candidate
    if current.parsed is None and candidate.parsed is None and not current.raw and candidate.raw:
        return candidate
    return current


def normalize_time_output(raw_value: object) -> TimeValue:
    parsed = parse_datetime(raw_value)
    if parsed is not None:
        return TimeValue(raw=parsed.strftime(DEFAULT_TIME_FORMAT), parsed=parsed)
    return TimeValue(raw=normalize_text(raw_value), parsed=None)


def build_catalog_time_map(
    rows: List[Dict[str, object]],
    sku_col: str,
    time_col: Optional[str],
) -> Dict[str, TimeValue]:
    catalog_map: Dict[str, TimeValue] = {}
    for row in rows:
        sku = normalize_code(row.get(sku_col))
        if should_exclude_code(sku):
            continue

        time_value = normalize_time_output(row.get(time_col)) if time_col else TimeValue(raw="", parsed=None)
        if sku not in catalog_map:
            catalog_map[sku] = time_value
        else:
            catalog_map[sku] = choose_latest_time(catalog_map[sku], time_value)

    return catalog_map


def build_shop_sku_map(rows: List[Dict[str, object]], shop_col: str, sku_col: str) -> Dict[str, Set[str]]:
    result: Dict[str, Set[str]] = {}
    for row in rows:
        shop = normalize_text(row.get(shop_col))
        if not shop:
            continue
        sku = normalize_code(row.get(sku_col))
        if should_exclude_code(sku):
            continue
        if shop not in result:
            result[shop] = set()
        result[shop].add(sku)
    return result


def result_sort_key(row: Dict[str, object]) -> Tuple[str, int, float, str]:
    parsed = row.get("_sort_time")
    if isinstance(parsed, datetime):
        return (str(row["店铺名称"]), 0, -parsed.timestamp(), str(row["商品编码"]))
    return (str(row["店铺名称"]), 1, 0.0, str(row["商品编码"]))


def build_result_rows(shop_sku_map: Dict[str, Set[str]], catalog_map: Dict[str, TimeValue]) -> List[Dict[str, object]]:
    result: List[Dict[str, object]] = []
    all_catalog_skus = set(catalog_map.keys())
    for shop in sorted(shop_sku_map.keys()):
        missing_skus = all_catalog_skus - shop_sku_map[shop]
        for sku in missing_skus:
            tv = catalog_map[sku]
            result.append(
                {
                    "店铺名称": shop,
                    "商品编码": sku,
                    "创建时间": tv.raw,
                    "_sort_time": tv.parsed,
                }
            )

    result.sort(key=result_sort_key)
    for row in result:
        row.pop("_sort_time", None)
    return result


def build_summary_rows(detail_rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    counter: Dict[str, int] = {}
    for row in detail_rows:
        shop = str(row["店铺名称"])
        counter[shop] = counter.get(shop, 0) + 1
    return [
        {"店铺名称": shop, "未出现商品数": count}
        for shop, count in sorted(counter.items(), key=lambda x: x[1], reverse=True)
    ]


def write_output(path: Path, detail_rows: List[Dict[str, object]], summary_rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "未出现明细"
    ws_detail.append(DETAIL_HEADERS)
    for row in detail_rows:
        ws_detail.append([row.get(h, "") for h in DETAIL_HEADERS])

    ws_summary = wb.create_sheet("店铺汇总")
    ws_summary.append(SUMMARY_HEADERS)
    for row in summary_rows:
        ws_summary.append([row.get(h, "") for h in SUMMARY_HEADERS])

    wb.save(path)


def choose_latest_input_file(input_dir: Path, prefix: str) -> Path:
    files = sorted(input_dir.glob(f"{prefix}*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"{input_dir} 下找不到以 `{prefix}` 开头的 xlsx 文件")
    return files[0]


def main() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    default_input_dir = base_dir / "data" / "input"
    default_output = base_dir / "data" / "output" / "店铺未上架商品编码.xlsx"

    parser = argparse.ArgumentParser(description="按店铺找出未出现在店铺商品资料中的商品编码（精确匹配）")
    parser.add_argument("--shop-file", default=None, help="店铺商品资料文件路径，不传则自动从 data/input 选择最新文件")
    parser.add_argument("--catalog-file", default=None, help="商品资料文件路径，不传则自动从 data/input 选择最新文件")
    parser.add_argument("--output", default=str(default_output), help="输出 xlsx 路径")
    parser.add_argument("--shop-sheet", default=None, help="店铺商品资料工作表名称，不传默认第一个")
    parser.add_argument("--catalog-sheet", default=None, help="商品资料工作表名称，不传默认第一个")
    parser.add_argument("--shop-col", default="店铺名称", help="店铺列名")
    parser.add_argument("--shop-sku-col", default="原始商品编码", help="店铺商品编码列名")
    parser.add_argument("--catalog-sku-col", default="商品编码", help="商品资料编码列名")
    parser.add_argument("--time-col", default="创建时间", help="商品资料时间列名")
    args = parser.parse_args()

    if args.shop_file:
        shop_file = Path(args.shop_file)
    else:
        shop_file = choose_latest_input_file(default_input_dir, "店铺商品资料")

    if args.catalog_file:
        catalog_file = Path(args.catalog_file)
    else:
        catalog_file = choose_latest_input_file(default_input_dir, "商品资料")

    output_path = Path(args.output)

    shop_columns, shop_rows = read_excel(shop_file, args.shop_sheet)
    catalog_columns, catalog_rows = read_excel(catalog_file, args.catalog_sheet)

    shop_col = resolve_column(shop_columns, args.shop_col)
    shop_sku_col = resolve_column(shop_columns, args.shop_sku_col)
    catalog_sku_col = resolve_column(catalog_columns, args.catalog_sku_col)
    time_col = resolve_column_or_none(catalog_columns, args.time_col)

    catalog_map = build_catalog_time_map(catalog_rows, catalog_sku_col, time_col)
    shop_sku_map = build_shop_sku_map(shop_rows, shop_col, shop_sku_col)
    detail_rows = build_result_rows(shop_sku_map, catalog_map)
    summary_rows = build_summary_rows(detail_rows)
    write_output(output_path, detail_rows, summary_rows)

    print(f"店铺商品资料: {shop_file}")
    print(f"商品资料: {catalog_file}")
    print(f"输出文件: {output_path}")
    print(f"店铺数: {len(shop_sku_map)}")
    print(f"未出现商品编码总数: {len(detail_rows)}")


if __name__ == "__main__":
    main()
