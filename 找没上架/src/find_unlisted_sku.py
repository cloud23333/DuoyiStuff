#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


@dataclass
class TimeValue:
    raw: str
    parsed: Optional[datetime]


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
DATETIME_FORMATS = (
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d",
    "%Y-%m-%d",
)
DETAIL_HEADERS = ["店铺", "未上架SKU", "创建时间"]
SUMMARY_HEADERS = ["店铺", "未上架SKU数量"]
CHINESE_CHAR_PATTERN = re.compile(r"[\u4e00-\u9fff]")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def should_exclude_sku(sku: str) -> bool:
    normalized = normalize_text(sku)
    if not normalized:
        return True
    if normalized.upper().startswith("JH"):
        return True
    if CHINESE_CHAR_PATTERN.search(normalized):
        return True
    return False


def parse_datetime(value: object) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = normalize_text(value)
    if not text:
        return None

    for fmt in DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def resolve_column(columns: Iterable[str], target: str) -> str:
    normalized_target = target.strip()
    for col in columns:
        if col.strip() == normalized_target:
            return col
    raise KeyError(f"找不到列: {target}")


def resolve_column_or_none(columns: Iterable[str], target: str) -> Optional[str]:
    try:
        return resolve_column(columns, target)
    except KeyError:
        return None


def read_csv(path: Path) -> Tuple[List[str], List[Dict[str, object]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"CSV 文件没有表头: {path}")
        columns = [str(c).strip() for c in reader.fieldnames]
        rows: List[Dict[str, object]] = []
        for row in reader:
            cleaned = {str(k).strip(): v for k, v in row.items()}
            rows.append(cleaned)
    return columns, rows


def read_excel(path: Path, sheet_name: Optional[str] = None) -> Tuple[List[str], List[Dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {sheet_name}")

    ws = wb[sheet_name]
    ws.reset_dimensions()
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration as e:
        raise ValueError(f"空表: {path} / {sheet_name}") from e

    columns = [normalize_text(c) for c in header_row]
    if not any(columns):
        raise ValueError(f"表头为空: {path} / {sheet_name}")

    rows: List[Dict[str, object]] = []
    for raw_row in row_iter:
        if raw_row is None:
            continue
        row_data: Dict[str, object] = {}
        has_value = False
        for col_name, cell in zip(columns, raw_row):
            if not col_name:
                continue
            row_data[col_name] = cell
            if normalize_text(cell):
                has_value = True
        if has_value:
            rows.append(row_data)

    return columns, rows


def read_table(path: Path, sheet_name: Optional[str] = None) -> Tuple[List[str], List[Dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        return read_excel(path, sheet_name=sheet_name)
    if suffix == ".csv":
        return read_csv(path)
    raise ValueError(f"暂不支持的文件类型: {path}")


def choose_latest_time(old: TimeValue, new: TimeValue) -> TimeValue:
    if old.parsed is None and new.parsed is not None:
        return new
    if old.parsed is not None and new.parsed is not None and new.parsed > old.parsed:
        return new
    if old.parsed is None and new.parsed is None and not old.raw and new.raw:
        return new
    return old


def build_catalog_map(
    rows: List[Dict[str, object]],
    sku_col: str,
    time_col: Optional[str],
) -> Dict[str, TimeValue]:
    result: Dict[str, TimeValue] = {}
    for row in rows:
        sku = normalize_text(row.get(sku_col, ""))
        if should_exclude_sku(sku):
            continue
        time_cell = row.get(time_col) if time_col else None
        tv = TimeValue(
            raw=normalize_text(time_cell) if time_col else "",
            parsed=parse_datetime(time_cell) if time_col else None,
        )
        if sku not in result:
            result[sku] = tv
        else:
            result[sku] = choose_latest_time(result[sku], tv)
    return result


def build_shop_listing(
    rows: List[Dict[str, object]],
    sku_col: str,
    shop_col: Optional[str],
    default_shop: str,
    status_col: Optional[str],
    down_status: str,
) -> Dict[str, set]:
    shop_to_skus: Dict[str, set] = defaultdict(set)
    normalized_down_status = normalize_text(down_status)
    for row in rows:
        sku = normalize_text(row.get(sku_col, ""))
        if should_exclude_sku(sku):
            continue

        if status_col:
            status = normalize_text(row.get(status_col, ""))
            if status == normalized_down_status:
                continue

        if shop_col:
            shop = normalize_text(row.get(shop_col, "")) or default_shop
        else:
            shop = default_shop
        shop_to_skus[shop].add(sku)

    if not shop_to_skus:
        shop_to_skus[default_shop] = set()
    return shop_to_skus


def build_missing_rows(
    shop_to_skus: Dict[str, set],
    catalog_map: Dict[str, TimeValue],
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    catalog_skus = set(catalog_map.keys())
    for shop in sorted(shop_to_skus.keys()):
        listed = shop_to_skus[shop]
        missing = catalog_skus - listed
        for sku in missing:
            tv = catalog_map[sku]
            display_time = tv.raw
            if tv.parsed is not None:
                display_time = tv.parsed.strftime("%Y-%m-%d %H:%M:%S")
            rows.append(
                {
                    "店铺": shop,
                    "未上架SKU": sku,
                    "创建时间": display_time,
                    "_sort_time": tv.parsed,
                }
            )

    rows.sort(key=missing_row_sort_key)

    for row in rows:
        row.pop("_sort_time", None)
    return rows


def missing_row_sort_key(row: Dict[str, object]) -> Tuple[str, int, float, str]:
    sort_time = row["_sort_time"]
    if isinstance(sort_time, datetime):
        time_bucket = 0
        time_rank = -sort_time.timestamp()
    else:
        time_bucket = 1
        time_rank = 0.0
    return (
        str(row["店铺"]),
        time_bucket,
        time_rank,
        str(row["未上架SKU"]),
    )


def build_summary(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    counter: Dict[str, int] = defaultdict(int)
    for row in rows:
        counter[str(row["店铺"])] += 1
    return [
        {"店铺": shop, "未上架SKU数量": count}
        for shop, count in sorted(counter.items(), key=lambda x: x[1], reverse=True)
    ]


def write_excel_output(path: Path, detail_rows: List[Dict[str, object]], summary_rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "未上架明细"
    ws_detail.append(DETAIL_HEADERS)
    for row in detail_rows:
        ws_detail.append([row.get(h, "") for h in DETAIL_HEADERS])

    ws_summary = wb.create_sheet("店铺汇总")
    ws_summary.append(SUMMARY_HEADERS)
    for row in summary_rows:
        ws_summary.append([row.get(h, "") for h in SUMMARY_HEADERS])

    wb.save(path)


def write_csv_output(path: Path, detail_rows: List[Dict[str, object]], summary_rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=DETAIL_HEADERS)
        writer.writeheader()
        writer.writerows(detail_rows)

    summary_path = path.with_name(f"{path.stem}_summary.csv")
    with summary_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_HEADERS)
        writer.writeheader()
        writer.writerows(summary_rows)


def resolve_shop_column(listed_cols: List[str], shop_col_name: str, allow_missing: bool) -> Optional[str]:
    shop_col = resolve_column_or_none(listed_cols, shop_col_name)
    if shop_col is not None:
        return shop_col
    if allow_missing:
        return None
    raise ValueError(
        f"已上架表缺少店铺列 `{shop_col_name}`。"
        f"当前列为: {listed_cols}。"
        "若你确认文件没有店铺列，请加 --allow-missing-shop-col。"
    )


def resolve_status_column(listed_cols: List[str], status_col_name: str, allow_missing: bool) -> Optional[str]:
    status_col = resolve_column_or_none(listed_cols, status_col_name)
    if status_col is not None:
        return status_col
    if allow_missing:
        return None
    raise ValueError(
        f"已上架表缺少状态列 `{status_col_name}`。"
        f"当前列为: {listed_cols}。"
        "若你确认文件没有状态列，请加 --allow-missing-status-col。"
    )


def write_output(path: Path, detail_rows: List[Dict[str, object]], summary_rows: List[Dict[str, object]]) -> None:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        write_excel_output(path, detail_rows, summary_rows)
    elif suffix == ".csv":
        write_csv_output(path, detail_rows, summary_rows)
    else:
        raise ValueError("输出文件仅支持 .xlsx 或 .csv")


def main() -> None:
    parser = argparse.ArgumentParser(description="按店铺找未上架 SKU，并按时间倒序")
    parser.add_argument("--listed-file", required=True, help="已上架 SKU 表路径")
    parser.add_argument("--catalog-file", required=True, help="全量商品表路径")
    parser.add_argument("--output", default="data/output/未上架SKU结果.xlsx", help="输出路径(.xlsx/.csv)")

    parser.add_argument("--listed-sheet", default=None, help="已上架表 sheet 名，不填默认第一个")
    parser.add_argument("--catalog-sheet", default=None, help="全量商品表 sheet 名，不填默认第一个")

    parser.add_argument("--listed-sku-col", default="平台商品基本信息-SKU货号", help="已上架表 SKU 列名")
    parser.add_argument("--catalog-sku-col", default="款式编码", help="全量商品表 SKU 列名")
    parser.add_argument("--time-col", default="创建时间", help="全量商品表时间列名")
    parser.add_argument("--shop-col", default="店铺", help="已上架表店铺列名（默认必须存在）")
    parser.add_argument("--status-col", default="平台商品基本信息-选品状态", help="已上架表选品状态列名")
    parser.add_argument("--down-status", default="已下架", help="判定为未上架的状态值")
    parser.add_argument("--default-shop", default="默认店铺", help="店铺为空或未提供店铺列时的回退值")
    parser.add_argument(
        "--allow-missing-shop-col",
        action="store_true",
        help="允许已上架表缺少店铺列；不传该参数时，缺少店铺列会直接报错",
    )
    parser.add_argument(
        "--allow-missing-status-col",
        action="store_true",
        help="允许已上架表缺少状态列；不传该参数时，缺少状态列会直接报错",
    )

    args = parser.parse_args()

    listed_path = Path(args.listed_file)
    catalog_path = Path(args.catalog_file)
    output_path = Path(args.output)

    listed_cols, listed_rows = read_table(listed_path, args.listed_sheet)
    catalog_cols, catalog_rows = read_table(catalog_path, args.catalog_sheet)

    listed_sku_col = resolve_column(listed_cols, args.listed_sku_col)
    catalog_sku_col = resolve_column(catalog_cols, args.catalog_sku_col)

    shop_col = resolve_shop_column(listed_cols, args.shop_col, args.allow_missing_shop_col)
    status_col = resolve_status_column(listed_cols, args.status_col, args.allow_missing_status_col)
    time_col = resolve_column_or_none(catalog_cols, args.time_col)

    catalog_map = build_catalog_map(catalog_rows, catalog_sku_col, time_col)
    shop_to_skus = build_shop_listing(
        listed_rows,
        listed_sku_col,
        shop_col,
        args.default_shop,
        status_col,
        args.down_status,
    )

    detail_rows = build_missing_rows(shop_to_skus, catalog_map)
    summary_rows = build_summary(detail_rows)

    write_output(output_path, detail_rows, summary_rows)

    print(f"输出完成: {output_path}")
    print(f"店铺数: {len(shop_to_skus)}")
    print(f"未上架SKU总数: {len(detail_rows)}")


if __name__ == "__main__":
    main()
